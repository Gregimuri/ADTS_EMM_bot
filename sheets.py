from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO, StringIO
import csv
import logging
from pathlib import Path

import httpx
from openpyxl import load_workbook

from config import COLUMNS, LOCAL_XLSX, SHEET_GID, SPREADSHEET_ID

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class PlayerRow:
    object_number: str
    address: str
    player_id: str
    name: str
    emm: str
    reflash: str
    model: str
    adb: str
    update_cube: str
    do_nothing: str
    old_version: str
    edtech: str
    mac: str


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() == "none":
        return ""
    return text


def _row_from_values(values: list[object] | tuple[object, ...]) -> PlayerRow | None:
    if len(values) < 4:
        return None
    padded = list(values) + [""] * (len(COLUMNS) - len(values))
    name = _cell(padded[3])
    if not name:
        return None
    return PlayerRow(
        object_number=_cell(padded[0]),
        address=_cell(padded[1]),
        player_id=_cell(padded[2]),
        name=name,
        emm=_cell(padded[4]),
        reflash=_cell(padded[5]),
        model=_cell(padded[6]),
        adb=_cell(padded[7]),
        update_cube=_cell(padded[8]),
        do_nothing=_cell(padded[9]),
        old_version=_cell(padded[10]),
        edtech=_cell(padded[11]),
        mac=_cell(padded[12]),
    )


def _parse_csv_text(text: str) -> list[PlayerRow]:
    reader = csv.reader(StringIO(text))
    rows: list[PlayerRow] = []
    header_skipped = False
    for raw in reader:
        if not raw:
            continue
        if not header_skipped:
            header_skipped = True
            # skip header if it looks like one
            first = _cell(raw[0]).lower()
            if first in {"objectnumber", "object_number"} or "адрес" in "".join(raw).lower():
                continue
        item = _row_from_values(raw)
        if item:
            rows.append(item)
    return rows


def _load_from_xlsx(path: Path) -> list[PlayerRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows: list[PlayerRow] = []
        for i, raw in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            item = _row_from_values(raw or ())
            if item:
                rows.append(item)
        return rows
    finally:
        wb.close()


def fetch_google_sheet_csv() -> list[PlayerRow]:
    url = (
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}"
        f"/export?format=csv&gid={SHEET_GID}"
    )
    with httpx.Client(follow_redirects=True, timeout=60.0) as client:
        response = client.get(url, headers={"User-Agent": "EMM-Telegram-Bot/1.0"})
        response.raise_for_status()
        content_type = response.headers.get("content-type", "")
        text = response.content.decode("utf-8-sig", errors="replace")
        if "text/html" in content_type or text.lstrip().lower().startswith("<!doctype"):
            raise PermissionError(
                "Google Таблица недоступна без авторизации. "
                "Откройте доступ: «Настройки доступа» → «Все, у кого есть ссылка» → «Читатель»."
            )
        rows = _parse_csv_text(text)
        if not rows:
            raise RuntimeError("CSV из Google Таблицы пустой или не распознан.")
        return rows


class SheetStore:
    def __init__(self) -> None:
        self.rows: list[PlayerRow] = []
        self.source: str = "empty"

    def reload(self) -> str:
        try:
            self.rows = fetch_google_sheet_csv()
            self.source = "google"
            logger.info("Loaded %s rows from Google Sheets", len(self.rows))
            return f"Загружено из Google Таблицы: {len(self.rows)} строк."
        except Exception as exc:
            logger.warning("Google Sheets load failed: %s", exc)
            if LOCAL_XLSX.exists():
                self.rows = _load_from_xlsx(LOCAL_XLSX)
                self.source = "xlsx"
                logger.info("Loaded %s rows from local xlsx", len(self.rows))
                return (
                    f"Google недоступна ({exc}). "
                    f"Запасной файл Excel: {len(self.rows)} строк."
                )
            self.rows = []
            self.source = "empty"
            raise

    def find_by_name(self, query: str) -> list[PlayerRow]:
        needle = query.casefold().strip()
        if not needle:
            return []
        return [row for row in self.rows if needle in row.name.casefold()]
